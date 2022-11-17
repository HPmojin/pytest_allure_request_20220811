#!/user/bin/env python3
# -*- coding: utf-8 -*-
#-------------------------------------------------------------------------------
# @File : operation_excle.py
# @Time : 2022-03-23 12:04
# @Author : mojin
# @Email : 397135766@qq.com
# @Software : PyCharm
#-------------------------------------------------------------------------------

from openpyxl.styles import Alignment
from openpyxl.styles import Border,Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font  # 导入字体模块
import openpyxl,string
from common.logger import Logger




class operation_excle:


    @classmethod
    def read_excel(cls,file_name, case_severity_list):


        wb = openpyxl.load_workbook(file_name)
        #ws = wb.active#打开当前页
        sheet_names = wb.sheetnames  # 得到工作簿的所有工作表名 结果： ['Sheet1', 'Sheet2', 'Sheet3']
        rows_list = []
        for title in sheet_names:
            ws = wb[title]  # 打开指定页
            # 取出每行的值，以list方式存放

            #rows_list=[[],[],[]]
            for row in ws.rows:
                row_list = []
                for cell in row:
                    if cell.value == None:
                        cell_srt = ''
                    else:
                        cell_srt = cell.value
                    #print(cell.value)
                    row_list.append(cell_srt)

                #row_list.append(title)
                row_list.insert(0,title)

                if row_list[5] in case_severity_list:#筛选匹配的用例等级进行测试

                    rows_list.append(row_list)
            #print(rows_list)
            # 结果转换成键值对的形式存放

        result = []
        for i in range(len(rows_list) - 1):
            row_dict = {}
            for j in range(len(rows_list[0])):
                row_dict[rows_list[0][j]] = rows_list[i + 1][j]
            result.append(row_dict)

        #return result#返回字典形式

        return rows_list  # 返回列表形式


    def str_count_to_width(self,str):
        '''找出字符串中的中英文、空格、数字、标点符号个数'''
        count_en = count_dg = count_sp = count_zh = count_pu = 0
        for s in str:
            # 英文
            if s in string.ascii_letters:
                count_en += 1
            # 数字
            elif s.isdigit():
                count_dg += 1
            # 空格
            elif s.isspace():
                count_sp += 1
            # 中文，除了英文之外，剩下的字符认为就是中文
            elif s.isalpha():
                count_zh += 1
            # 特殊字符
            else:
                count_pu += 1

        # print('英文字符：', count_en)
        # print('数字：', count_dg)
        # print('空格：', count_sp)
        # print('中文字符：', count_zh)
        # print('特殊字符：', count_pu)
        #width = count_en * 1.8 + count_dg * 1.8 + count_sp * 1 + count_zh * 2.6 + count_pu * 2
        width = count_en * 1 + count_dg * 1.5 + count_sp * 1 + count_zh * 2 + count_pu * 1.25#宋体 11号的参数，其他格式可自己尝试
        return (float('%.2f' % width))

    # 数据写入Excel
    @classmethod
    def write_to_excel(cls,path: str, sheetStr, info, data):  # 数据写入Excel
        height, horizontal = (36, 'left')  # 宽度，高度 居中
        #     实例化一个workbook对象
        workbook = openpyxl.Workbook()
        ws = workbook[workbook.sheetnames[0]]
        # 激活一个sheet
        sheet = workbook.active
        # 为sheet设置一个title
        sheet.title = sheetStr

        # 添加表头（不需要表头可以不用加）
        data.insert(0, list(info))
        # 开始遍历数组


        width_list_all=[9] *len(data[0])#创建一个列表初始值为9的列表,表格的最小宽度

        for row_index, row_item in enumerate(data):
            #print(row_index,row_item)

            # 设置边框
            border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))  # 设置边框
            # 设置字体样式
            font = Font(u'宋体', size=11, bold=False, italic=False, strike=False, color='000000', )  # 设置字体样式
            # 设置居中
            align = Alignment(horizontal=horizontal, vertical='center', wrap_text=True)  # 设置居中

            for col_index, col_item in enumerate(row_item):
                #计算宽度后放到列表比较大小，去最大值生成新列表
                width_dg=self.str_count_to_width(col_item)
                if width_list_all[col_index]<width_dg:
                    width_list_all[col_index]=width_dg

                # 写入
                sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item).alignment = align  # 设置居中
                sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item).font = font  # 序列
                sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item).border = border  # 设置边框

        #print(width_list_all)
        ###修改行距 列距
        # width = 16
        # height = 36
        #print("row:", ws.max_row, "column:", ws.max_column)
        #设置高
        for i in range(1, ws.max_row + 1):
            ws.row_dimensions[i].height = height
        #设置宽
        for i in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = width_list_all[i-1]

        # 写入excel文件 如果path路径的文件不存在那么就会自动创建
        workbook.save(path)
        #print('写入成功')



